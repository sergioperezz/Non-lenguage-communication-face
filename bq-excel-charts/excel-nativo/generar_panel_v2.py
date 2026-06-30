"""Genera el panel GENÉRICO (Fase 1, v2) — versión ampliada.

Motor guiado por datos: NO se pre-genera un gráfico por caso. Una tabla de datos
en formato largo + SEIS parámetros que el usuario combina libremente. Cualquier
combinación (métrica × dimensión × periodo × comparación × tipo de gráfico) se
construye al vuelo con fórmulas + macro.

Parámetros (celdas del Panel):
  B3 Tipo de activo   (RF / RV)                          -> filtro
  B4 Métrica          (cascada según B3)                 -> qué se mide
  B5 Dimensión/eje X  (tiempo o composición, cascada)    -> cómo se desglosa
  B6 Periodo          (MTD/YTD/3M/6M/1A/3A)              -> ventana temporal
  B7 Comparación      (Cartera+Benchmark/Solo...)        -> serie(s)
  B8 Tipo de gráfico  (Columnas/Barras/Líneas/...)       -> representación

Datos de ejemplo de COBERTURA COMPLETA: hay valores ficticios para todas las
combinaciones seleccionables, así que en "modo diseño" cualquier gráfico se ve.
"""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


# === Modelo de ejemplo ======================================================
METRICAS = {
    "RF": ["Rentabilidad", "TIR", "TER", "Duración", "Spread", "Liquidez", "Peso"],
    "RV": ["Rentabilidad", "TER", "PER", "DividendYield", "Liquidez", "Peso"],
}

# Dimensiones (eje X) por tipo de activo: de tiempo + de composición.
# (Tokens SIN acentos: se usan para construir nombres definidos vía INDIRECT.)
DIM_TIEMPO = ["Mensual", "Trimestral", "Semestral", "Anual"]
DIM_COMP_COMUN = ["Activo", "Geografia", "Industria", "Sector", "Divisa"]
EJES = {
    "RF": DIM_TIEMPO + DIM_COMP_COMUN + ["Rating"],   # Rating solo RF
    "RV": DIM_TIEMPO + DIM_COMP_COMUN,
}


def _meses(n: int, y0: int, m0: int) -> list[str]:
    out, y, m = [], y0, m0
    for _ in range(n):
        out.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


CATEGORIAS = {
    # Tiempo (3 años hasta 2026-06)
    "Mensual": _meses(36, 2023, 7),                                   # 36
    "Trimestral": (
        ["2023-T3", "2023-T4"]
        + [f"{y}-T{q}" for y in (2024, 2025) for q in (1, 2, 3, 4)]
        + ["2026-T1", "2026-T2"]
    ),                                                                # 12
    "Semestral": ["2023-S2", "2024-S1", "2024-S2", "2025-S1", "2025-S2", "2026-S1"],  # 6
    "Anual": ["2024", "2025", "2026"],                               # 3
    # Composición
    "Activo": ["Deuda Pública", "Deuda Privada", "Acciones", "Liquidez", "Derivados"],  # 5
    "Geografia": ["Europa", "Norteamérica", "Asia-Pacífico", "Emergentes", "Latam"],    # 5
    "Industria": ["Financiero", "Tecnología", "Salud", "Energía", "Consumo",
                  "Industrial", "Utilities"],                         # 7
    "Sector": ["Financiero", "Industrial", "Tecnología", "Consumo", "Energía", "Salud"],  # 6
    "Divisa": ["EUR", "USD", "GBP", "JPY", "CHF", "Otras"],          # 6
    "Rating": ["AAA", "AA", "A", "BBB", "BB", "B"],                  # 6
}
MAX_CATS = max(len(c) for c in CATEGORIAS.values())  # filas de la tabla de resultados

# (base, paso, factor del benchmark) por métrica.
PARAMS = {
    "Rentabilidad": (9.0, 0.60, 0.90),
    "TIR": (3.5, 0.20, 1.05),
    "TER": (0.85, 0.03, 1.00),
    "Duración": (5.0, 0.30, 0.92),
    "Spread": (120.0, 15.0, 0.85),
    "Liquidez": (5.0, 0.10, 1.02),
    "Peso": (18.0, 3.0, 1.00),
    "PER": (15.0, 0.80, 1.04),
    "DividendYield": (2.8, 0.20, 0.95),
}

# Periodo -> nº de buckets recientes a mostrar, por granularidad temporal.
PERIODOS = ["MTD", "YTD", "3M", "6M", "1A", "3A"]
TABLA_N = {                       # filas=Periodos, cols=DIM_TIEMPO (Mensual..Anual)
    "MTD": [1, 1, 1, 1],
    "YTD": [6, 2, 1, 1],
    "3M": [3, 1, 1, 1],
    "6M": [6, 2, 1, 1],
    "1A": [12, 4, 2, 1],
    "3A": [36, 12, 6, 3],
}

AZUL = "FF0072CE"
GRIS = "FFF2F2F2"
BOLD = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFFFF")


def _valor(metrica: str, idx: int) -> float:
    """Valor ficticio plausible: oscila alrededor de la base."""
    base, paso, _ = PARAMS[metrica]
    return round(base + paso * (2.0 * math.sin(idx * 0.6) + (idx % 4) - 1.5), 2)


def build() -> Workbook:
    wb = Workbook()

    # === Datos: TipoActivo | Metrica | EjeXTipo | EjeXValor | Serie | Valor ===
    ws_d = wb.active
    ws_d.title = "Datos"
    ws_d.append(["TipoActivo", "Metrica", "EjeXTipo", "EjeXValor", "Serie", "Valor"])
    for c in ws_d[1]:
        c.font = BOLD
    for tipo, metricas in METRICAS.items():
        for met in metricas:
            factor = PARAMS[met][2]
            for ejx in EJES[tipo]:
                for idx, cat in enumerate(CATEGORIAS[ejx]):
                    cartera = _valor(met, idx)
                    ws_d.append([tipo, met, ejx, cat, "Cartera", cartera])
                    ws_d.append([tipo, met, ejx, cat, "Benchmark", round(cartera * factor, 2)])
    n = ws_d.max_row  # última fila de datos (para acotar SUMIFS = rápido)

    # === Listas: cascadas, categorías por dimensión y tabla de periodos ===
    ws_l = wb.create_sheet("Listas")

    def col(letter: str, header: str, items: list) -> str:
        ws_l[f"{letter}1"] = header
        ws_l[f"{letter}1"].font = BOLD
        for i, v in enumerate(items, start=2):
            ws_l[f"{letter}{i}"] = v
        return f"Listas!${letter}$2:${letter}${1 + len(items)}"

    def name(nm: str, ref: str) -> None:
        wb.defined_names.add(DefinedName(nm, attr_text=ref))

    name("RF", col("A", "RF", METRICAS["RF"]))
    name("RV", col("B", "RV", METRICAS["RV"]))
    name("Eje_RF", col("C", "Eje_RF", EJES["RF"]))
    name("Eje_RV", col("D", "Eje_RV", EJES["RV"]))

    cat_cols = ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    for letter, dim in zip(cat_cols, CATEGORIAS):
        name(f"Cat_{dim}", col(letter, f"Cat_{dim}", CATEGORIAS[dim]))

    name("Periodos", col("P", "Periodos", PERIODOS))
    name("DimTiempo", col("Q", "DimTiempo", DIM_TIEMPO))
    # Matriz N (R2:U7): filas=Periodos, cols=DIM_TIEMPO
    ws_l["R1"], ws_l["S1"], ws_l["T1"], ws_l["U1"] = ("N_Men", "N_Tri", "N_Sem", "N_Anu")
    for r, p in enumerate(PERIODOS, start=2):
        for cidx, val in enumerate(TABLA_N[p]):
            ws_l.cell(row=r, column=18 + cidx, value=val)  # 18 = R
    name("TablaN", f"Listas!$R$2:$U${1 + len(PERIODOS)}")

    # === Panel ===
    ws = wb.create_sheet("Panel", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Panel genérico de gráficos"
    ws["A1"].font = Font(bold=True, size=14)

    controles = {
        "A3": "Tipo de activo",
        "A4": "Métrica",
        "A5": "Dimensión (eje X)",
        "A6": "Periodo",
        "A7": "Comparación",
        "A8": "Tipo de gráfico",
    }
    for celda, txt in controles.items():
        ws[celda] = txt
        ws[celda].font = BOLD
    ws["B3"], ws["B4"], ws["B5"] = "RF", "Rentabilidad", "Mensual"
    ws["B6"], ws["B7"], ws["B8"] = "1A", "Cartera + Benchmark", "Columnas"
    for celda in ("B3", "B4", "B5", "B6", "B7", "B8"):
        ws[celda].fill = PatternFill("solid", fgColor=GRIS)

    dvs = [
        ("B3", '"RF,RV"'),
        ("B4", "=INDIRECT($B$3)"),
        ("B5", '=INDIRECT("Eje_"&$B$3)'),
        ("B6", '"MTD,YTD,3M,6M,1A,3A"'),
        ("B7", '"Cartera + Benchmark,Solo cartera,Solo benchmark"'),
        ("B8", '"Columnas,Barras,Líneas,Área,Circular,Anillo,Radar"'),
    ]
    for celda, formula in dvs:
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[celda])

    # Helper: nº de buckets a mostrar. Para dimensiones de tiempo lo da la tabla
    # Periodo×Granularidad; para composición (MATCH falla) -> todas las categorías.
    ws["A10"] = "Buckets visibles"
    ws["A10"].font = Font(italic=True, size=9)
    ws["B10"] = (
        '=IFERROR(INDEX(TablaN,MATCH($B$6,Periodos,0),MATCH($B$5,DimTiempo,0)),'
        'COUNTA(INDIRECT("Cat_"&$B$5)))'
    )

    # Título dinámico
    ws.merge_cells("A12:C12")
    ws["A12"] = '="Métrica: "&B4&"  ·  Por: "&B5&"  ·  Periodo: "&B6&"  ·  "&B7'
    ws["A12"].font = Font(bold=True, size=12, color=AZUL[2:])
    ws["A12"].alignment = Alignment(horizontal="left")

    # Tabla de resultados: Categoría | Cartera | Benchmark (al vuelo)
    ws["D2"], ws["E2"], ws["F2"] = "Categoría", "Cartera", "Benchmark"
    for celda in ("D2", "E2", "F2"):
        ws[celda].font = BOLD_WHITE
        ws[celda].fill = PatternFill("solid", fgColor=AZUL)

    for i in range(MAX_CATS):
        r = 3 + i
        # Categoría i-ésima de las ÚLTIMAS N (B10) de la dimensión elegida.
        ws.cell(row=r, column=4, value=(
            f'=IF((ROW()-2)>$B$10,"",'
            f'INDEX(INDIRECT("Cat_"&$B$5),'
            f'COUNTA(INDIRECT("Cat_"&$B$5))-$B$10+(ROW()-2)))'
        ))
        # Cartera / Benchmark con SUMIFS acotado (Datos!$X$2:$X$n)
        ws.cell(row=r, column=5, value=(
            f'=IF(OR($B$7="Solo benchmark",$D{r}=""),"",'
            f'SUMIFS(Datos!$F$2:$F${n},Datos!$A$2:$A${n},$B$3,Datos!$B$2:$B${n},$B$4,'
            f'Datos!$C$2:$C${n},$B$5,Datos!$D$2:$D${n},$D{r},Datos!$E$2:$E${n},"Cartera"))'
        ))
        ws.cell(row=r, column=6, value=(
            f'=IF(OR($B$7="Solo cartera",$D{r}=""),"",'
            f'SUMIFS(Datos!$F$2:$F${n},Datos!$A$2:$A${n},$B$3,Datos!$B$2:$B${n},$B$4,'
            f'Datos!$C$2:$C${n},$B$5,Datos!$D$2:$D${n},$D{r},Datos!$E$2:$E${n},"Benchmark"))'
        ))

    # Gráfico enlazado a la tabla (la macro ajusta tipo y combo barras+línea)
    chart = BarChart()
    chart.type = "col"
    chart.title = "Cartera vs Benchmark"
    chart.height = 8.5
    chart.width = 18
    last = 2 + MAX_CATS
    data = Reference(ws, min_col=5, max_col=6, min_row=2, max_row=last)
    cats = Reference(ws, min_col=4, min_row=3, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "H2")

    for c, w in {"A": 18, "B": 22, "D": 16, "E": 12, "F": 12}.items():
        ws.column_dimensions[c].width = w

    return wb


if __name__ == "__main__":
    out = Path(__file__).parent / "Panel_Generico.xlsx"
    build().save(out)
    print(f"Generado: {out}  (máx categorías: {MAX_CATS})")
