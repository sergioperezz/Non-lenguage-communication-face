"""Genera el panel GENÉRICO (Fase 1) — con comparación de varias entidades.

Motor guiado por datos. Jerarquía:

  Tipo de entidad   -> Fondo / Cartera / Indice
  Entidad 1         -> cascada del tipo (principal)
  Entidad 2, 3      -> opcionales, para comparar varias a la vez
  Grupo de métrica  -> familia; Métrica -> cascada del grupo
  Dimensión (eje X) -> tiempo o composición
  Filtro Tipo activo-> Todos/RF/RV
  Periodo           -> MTD/YTD/1M..6M/1A..6A
  Benchmark         -> Con/Sin (de la Entidad 1) · Estilo Barras/Líneas
  Tipo de gráfico   -> representación

Cada entidad seleccionada es una serie -> permite "cartera A vs cartera B".
Datos de ejemplo (ficticios). En la Fase 2 vienen de BigQuery por ODBC.
"""

from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


# === Modelo ================================================================
ENTIDADES = {
    "Fondo": {
        "RF Privada A": ["RF"],
        "Bolsa Europa": ["RV"],
        "Mixto Moderado": ["RF", "RV"],
    },
    "Cartera": {
        "Cartera RF Gobierno": ["RF"],
        "Cartera RF Crédito": ["RF"],
        "Cartera Conservadora": ["RF", "RV"],
        "Cartera Dinamica": ["RF", "RV"],
    },
    "Indice": {
        "Euro Stoxx 50": ["RV"],
        "Bloomberg Agg": ["RF"],
        "MSCI World": ["RV"],
    },
}

GRUPOS = {
    "Rendimiento": ["Rentabilidad", "Rentab. acum."],
    "Riesgo": ["Duración", "TIR", "Spread", "Volatilidad", "Beta"],
    "Composicion": ["Peso"],
    "Costes": ["TER"],
    "Liquidez": ["Liquidez"],
    "Valoracion": ["PER", "DividendYield"],
}
ALL_METRICS = [m for ms in GRUPOS.values() for m in ms]

DIM_TIEMPO = ["Mensual", "Trimestral", "Semestral", "Anual"]
DIM_COMP = ["Activo", "Geografia", "Industria", "Sector", "Divisa", "Rating"]
DIMS = DIM_TIEMPO + DIM_COMP


def _meses(n, y0, m0):
    out, y, m = [], y0, m0
    for _ in range(n):
        out.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return out


CATEGORIAS = {
    "Mensual": _meses(36, 2023, 7),
    "Trimestral": (["2023-T3", "2023-T4"]
                   + [f"{y}-T{q}" for y in (2024, 2025) for q in (1, 2, 3, 4)]
                   + ["2026-T1", "2026-T2"]),
    "Semestral": ["2023-S2", "2024-S1", "2024-S2", "2025-S1", "2025-S2", "2026-S1"],
    "Anual": ["2024", "2025", "2026"],
    "Activo": ["Deuda Pública", "Deuda Privada", "Acciones", "Liquidez", "Derivados"],
    "Geografia": ["Europa", "Norteamérica", "Asia-Pacífico", "Emergentes", "Latam"],
    "Industria": ["Financiero", "Tecnología", "Salud", "Energía", "Consumo",
                  "Industrial", "Utilities"],
    "Sector": ["Financiero", "Industrial", "Tecnología", "Consumo", "Energía", "Salud"],
    "Divisa": ["EUR", "USD", "GBP", "JPY", "CHF", "Otras"],
    "Rating": ["AAA", "AA", "A", "BBB", "BB", "B"],
}
MAX_CATS = max(len(c) for c in CATEGORIAS.values())

PARAMS = {
    "Rentabilidad": (9.0, 0.60, 0.90), "Rentab. acum.": (12.0, 1.20, 0.90),
    "Duración": (5.0, 0.30, 0.92), "TIR": (3.5, 0.20, 1.05), "Spread": (120.0, 15.0, 0.85),
    "Volatilidad": (12.0, 0.80, 1.10), "Beta": (1.0, 0.05, 1.00), "Peso": (18.0, 3.0, 1.00),
    "TER": (0.85, 0.03, 1.00), "Liquidez": (5.0, 0.10, 1.02),
    "PER": (15.0, 0.80, 1.04), "DividendYield": (2.8, 0.20, 0.95),
}

PERIODOS = ["MTD", "YTD", "1M", "2M", "3M", "4M", "5M", "6M",
            "1A", "2A", "3A", "4A", "5A", "6A"]
_MESES_PERIODO = {"MTD": 1, "YTD": 6, "1M": 1, "2M": 2, "3M": 3, "4M": 4, "5M": 5,
                  "6M": 6, "1A": 12, "2A": 24, "3A": 36, "4A": 48, "5A": 60, "6A": 72}
_COUNTS = [36, 12, 6, 3]


def _n_row(meses):
    raw = [meses, math.ceil(meses / 3), math.ceil(meses / 6), math.ceil(meses / 12)]
    return [min(r, c) for r, c in zip(raw, _COUNTS)]


TABLA_N = {p: _n_row(_MESES_PERIODO[p]) for p in PERIODOS}

# Índice de la entidad (columna de datos) y offset por serie/entidad para variar
# los valores ficticios entre entidades (para que la comparación no salga igual).
_ENT_LIST = [e for ents in ENTIDADES.values() for e in ents]

AZUL, GRIS = "FF0072CE", "FFF2F2F2"
BOLD = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFFFF")


def _valor(metrica, idx, ent_off=0):
    base, paso, _ = PARAMS[metrica]
    return round(base + paso * (2.0 * math.sin((idx + ent_off) * 0.6) + (idx % 4) - 1.5)
                 + ent_off * paso * 0.4, 2)


def build() -> Workbook:
    wb = Workbook()

    # === Datos: Entidad|TipoActivo|Metrica|EjeXTipo|EjeXValor|Serie|Valor ===
    ws_d = wb.active
    ws_d.title = "Datos"
    ws_d.append(["Entidad", "TipoActivo", "Metrica", "EjeXTipo", "EjeXValor", "Serie", "Valor"])
    for c in ws_d[1]:
        c.font = BOLD
    for tipo_ent, entidades in ENTIDADES.items():
        for entidad, tipos in entidades.items():
            ent_off = _ENT_LIST.index(entidad)              # varía entre entidades
            for tipo in tipos:
                for met in ALL_METRICS:
                    factor = PARAMS[met][2]
                    for dim in DIMS:
                        for idx, cat in enumerate(CATEGORIAS[dim]):
                            v = _valor(met, idx, ent_off)
                            ws_d.append([entidad, tipo, met, dim, cat, "Cartera", v])
                            ws_d.append([entidad, tipo, met, dim, cat, "Benchmark", round(v * factor, 2)])
    n = ws_d.max_row

    # === Listas ===
    ws_l = wb.create_sheet("Listas")

    def col(letter, header, items):
        ws_l[f"{letter}1"] = header
        ws_l[f"{letter}1"].font = BOLD
        for i, v in enumerate(items, start=2):
            ws_l[f"{letter}{i}"] = v
        return f"Listas!${letter}$2:${letter}${1 + len(items)}"

    def name(nm, ref):
        wb.defined_names.add(DefinedName(nm, attr_text=ref))

    for tipo_ent, letter in {"Fondo": "A", "Cartera": "B", "Indice": "C"}.items():
        name(f"Ent_{tipo_ent}", col(letter, f"Ent_{tipo_ent}", list(ENTIDADES[tipo_ent])))

    grp_cols = ["D", "E", "F", "G", "H", "I"]
    for letter, g in zip(grp_cols, GRUPOS):
        name(f"Grupo_{g}", col(letter, f"Grupo_{g}", GRUPOS[g]))

    cat_cols = ["K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]
    for letter, dim in zip(cat_cols, CATEGORIAS):
        name(f"Cat_{dim}", col(letter, f"Cat_{dim}", CATEGORIAS[dim]))

    name("Periodos", col("V", "Periodos", PERIODOS))
    name("DimTiempo", col("W", "DimTiempo", DIM_TIEMPO))
    ws_l["Y1"], ws_l["Z1"], ws_l["AA1"], ws_l["AB1"] = ("N_Men", "N_Tri", "N_Sem", "N_Anu")
    for r, p in enumerate(PERIODOS, start=2):
        for cidx, val in enumerate(TABLA_N[p]):
            ws_l.cell(row=r, column=25 + cidx, value=val)  # 25 = Y
    name("TablaN", f"Listas!$Y$2:$AB${1 + len(PERIODOS)}")

    # === Panel ===
    ws = wb.create_sheet("Panel", 0)
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Panel genérico de gráficos"
    ws["A1"].font = Font(bold=True, size=14)

    etiquetas = {
        "A3": "Tipo de entidad", "A4": "Entidad 1", "A5": "Entidad 2 (opc.)",
        "A6": "Entidad 3 (opc.)", "A7": "Grupo de métrica", "A8": "Métrica",
        "A9": "Dimensión (eje X)", "A10": "Filtro: tipo de activo", "A11": "Periodo",
        "A12": "Benchmark", "A13": "Estilo benchmark", "A14": "Tipo de gráfico",
    }
    for celda, txt in etiquetas.items():
        ws[celda] = txt
        ws[celda].font = BOLD
    defaults = {
        "B3": "Fondo", "B4": "RF Privada A", "B5": "", "B6": "", "B7": "Riesgo",
        "B8": "Duración", "B9": "Trimestral", "B10": "Todos", "B11": "3A",
        "B12": "Con benchmark", "B13": "Líneas", "B14": "Columnas",
    }
    for celda, val in defaults.items():
        ws[celda] = val
        ws[celda].fill = PatternFill("solid", fgColor=GRIS)

    grupos_lst = ",".join(GRUPOS)
    dims_lst = ",".join(DIMS)
    periodos_lst = ",".join(PERIODOS)
    dvs = [
        ("B3", '"Fondo,Cartera,Indice"', False),
        ("B4", '=INDIRECT("Ent_"&$B$3)', False),
        ("B5", '=INDIRECT("Ent_"&$B$3)', True),      # opcional (puede quedar vacío)
        ("B6", '=INDIRECT("Ent_"&$B$3)', True),      # opcional
        ("B7", f'"{grupos_lst}"', False),
        ("B8", '=INDIRECT("Grupo_"&$B$7)', False),
        ("B9", f'"{dims_lst}"', False),
        ("B10", '"Todos,RF,RV"', False),
        ("B11", f'"{periodos_lst}"', False),
        ("B12", '"Con benchmark,Sin benchmark"', False),
        ("B13", '"Barras,Líneas"', False),
        ("B14", '"Columnas,Barras,Líneas,Área,Circular,Anillo,Radar"', False),
    ]
    for celda, formula, blank in dvs:
        dv = DataValidation(type="list", formula1=formula, allow_blank=blank)
        ws.add_data_validation(dv)
        dv.add(ws[celda])

    # Helpers
    ws["A16"] = "Buckets visibles"
    ws["A16"].font = Font(italic=True, size=9)
    ws["B16"] = ('=IFERROR(INDEX(TablaN,MATCH($B$11,Periodos,0),MATCH($B$9,DimTiempo,0)),'
                 'COUNTA(INDIRECT("Cat_"&$B$9)))')
    ws["A17"] = "Criterio tipo activo"
    ws["A17"].font = Font(italic=True, size=9)
    ws["B17"] = '=IF($B$10="Todos","*",$B$10)'

    # Título dinámico
    ws.merge_cells("A19:C19")
    ws["A19"] = ('="Entidad: "&B4&IF(B5<>""," vs "&B5,"")&IF(B6<>""," vs "&B6,"")'
                 '&"  ·  "&B8&" por "&B9&"  ·  "&B11')
    ws["A19"].font = Font(bold=True, size=12, color=AZUL[2:])
    ws["A19"].alignment = Alignment(horizontal="left")

    # Tabla de resultados: Categoría | Ent1 | Ent2 | Ent3 | Benchmark
    ws["D2"] = "Categoría"
    ws["E2"], ws["F2"], ws["G2"], ws["H2"] = "=B4", "=B5", "=B6", "Benchmark"
    for celda in ("D2", "E2", "F2", "G2", "H2"):
        ws[celda].font = BOLD_WHITE
        ws[celda].fill = PatternFill("solid", fgColor=AZUL)

    def sumifs(ent, serie):
        return (f'SUMIFS(Datos!$G$2:$G${n},Datos!$A$2:$A${n},{ent},'
                f'Datos!$B$2:$B${n},$B$17,Datos!$C$2:$C${n},$B$8,'
                f'Datos!$D$2:$D${n},$B$9,Datos!$E$2:$E${n},$D{{r}},'
                f'Datos!$F$2:$F${n},"{serie}")')

    for i in range(MAX_CATS):
        r = 3 + i
        ws.cell(row=r, column=4, value=(
            f'=IF((ROW()-2)>$B$16,"",INDEX(INDIRECT("Cat_"&$B$9),'
            f'COUNTA(INDIRECT("Cat_"&$B$9))-$B$16+(ROW()-2)))'
        ))
        ws.cell(row=r, column=5, value=(
            f'=IF(OR($B$4="",$D{r}=""),"",{sumifs("$B$4", "Cartera").format(r=r)})'))
        ws.cell(row=r, column=6, value=(
            f'=IF(OR($B$5="",$D{r}=""),"",{sumifs("$B$5", "Cartera").format(r=r)})'))
        ws.cell(row=r, column=7, value=(
            f'=IF(OR($B$6="",$D{r}=""),"",{sumifs("$B$6", "Cartera").format(r=r)})'))
        ws.cell(row=r, column=8, value=(
            f'=IF(OR($B$12="Sin benchmark",$B$4="",$D{r}=""),"",'
            f'{sumifs("$B$4", "Benchmark").format(r=r)})'))
        for cc in (5, 6, 7, 8):
            ws.cell(row=r, column=cc).number_format = "#,##0.00"

    # Gráfico inicial (Ent1 + Benchmark; la macro lo hace dinámico y multi-serie)
    vis0 = TABLA_N["3A"][DIM_TIEMPO.index("Trimestral")]
    last = 2 + vis0
    chart = BarChart()
    chart.type = "col"
    chart.title = "Comparación"
    chart.height = 8.5
    chart.width = 18
    chart.y_axis.title = "Duración"
    chart.x_axis.title = "Trimestral"
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.add_data(Reference(ws, min_col=5, min_row=2, max_row=last), titles_from_data=True)  # Ent1
    chart.add_data(Reference(ws, min_col=8, min_row=2, max_row=last), titles_from_data=True)  # Benchmark
    chart.set_categories(Reference(ws, min_col=4, min_row=3, max_row=last))
    ws.add_chart(chart, "H2")

    for c, w in {"A": 20, "B": 22, "D": 16, "E": 14, "F": 14, "G": 14, "H": 12}.items():
        ws.column_dimensions[c].width = w

    return wb


if __name__ == "__main__":
    out = Path(__file__).parent / "Panel_Generico.xlsx"
    build().save(out)
    print(f"Generado: {out}  (máx categorías: {MAX_CATS})")
